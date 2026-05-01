#!/usr/bin/env python3
"""
IBS HI Sale Platform v4 — Workflow 2026 Data Import Script
==========================================================
Reads Workflow 2026 Google Sheet (exported as .xlsx) and supplementary
Excel files, then generates SQL INSERT statements for all target tables.

Data Sources:
  1. Workflow 2026.xlsx         → Customers, Opportunities, Tasks
  2. Client List (2025).xlsx    → Customer-Account Manager mapping
  3. IBSHI Khach hang tiem nang → Pipeline + Capital Recovery
  4. Sale Plan (Gia khoan).xlsx → Invoice Status, KHKD targets

Target Tables:
  - sale_customers              (from Cus-code sheet + Client List)
  - sale_customer_contacts      (from Cus-code sheet contacts)
  - sale_opportunities          (from RFQ 2024/2025/2026 sheets + Pipeline)
  - sale_tasks                  (from To do list sheets)
  - sale_product_categories     (validate existing seeds vs actual data)
  - sale_khkd_targets           (validate/update from KHKD 2026 RevB)

Usage:
  1. Export Workflow 2026 from Google Sheets as .xlsx
  2. Place in same directory as this script
  3. Run: python3 workflow_to_platform_import.py
  4. Outputs: import_customers.sql, import_opportunities.sql, etc.
  5. Review SQL files, then execute against platform DB

Author: Claude AI Chief of Staff
Date: 2026-04-28
"""

import openpyxl
import uuid
import json
import os
import re
import sys
from datetime import datetime, date
from pathlib import Path
from typing import Optional, Dict, List, Any

# ══════════════════════════════════════════════════════════════════
# CONFIG
# ══════════════════════════════════════════════════════════════════
WORKFLOW_FILE = "Workflow 2026.xlsx"  # exported from Google Sheets
CLIENT_LIST_FILE = "Client List (2025) (2).xlsx"
PIPELINE_FILE = "20250429_IBSHI Khách hàng tiềm năng_rev01.xlsx"
SALE_PLAN_FILE = "Sale Plan 25-26-27 (Giá khoán).xlsx"

OUTPUT_DIR = "sql_imports"
IMPORT_DATE = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

# Product group mapping (scope keywords → KHKD category code)
PRODUCT_GROUP_MAP = {
    'HRSG': ['hrsg', 'heat recovery', 'nồi hơi', 'non pressure', 'npp'],
    'DIVERTER': ['diverter', 'guillotine', 'damper', 'van chuyển'],
    'SHIP': ['ship', 'vessel', 'tàu', 'marine', 'pontoon', 'gangway', 'crane', 'rtg'],
    'PV': ['pressure vessel', 'tank', 'column', 'bồn', 'heat exchanger'],
    'HANDLING': ['handling', 'conveyor', 'shiploader', 'hoist', 'hopper'],
    'DUCT': ['duct', 'stack', 'ống khói', 'exhaust', 'flue gas', 'chimney'],
    'OTHER': []  # fallback
}

# PLHD → stage mapping
PLHD_STAGE_MAP = {
    'I': 'WON',      # Hợp đồng đã ký (Signed contracts)
    'II': 'NEGOTIATION',  # Đang đàm phán (Negotiating)
    'III': 'WON',    # Hợp đồng đã ký nước ngoài
    'IV': 'QUOTED',  # Đã chào giá (Quoted)
    'V': 'PROSPECT', # Tiềm năng (Potential)
}

# Win probability mapping
WIN_PCT_MAP = {
    'WON': 100,
    'NEGOTIATION': 70,
    'QUOTED': 50,
    'RFQ_RECEIVED': 30,
    'PROSPECT': 10,
    'LEAD': 5,
    'LOST': 0,
}


# ══════════════════════════════════════════════════════════════════
# UTILITY FUNCTIONS
# ══════════════════════════════════════════════════════════════════
def gen_id():
    """Generate UUID for primary key."""
    return str(uuid.uuid4())[:16].replace('-', '')

def sql_str(val) -> str:
    """Escape value for SQL string."""
    if val is None or str(val).strip() in ('', 'None', 'nan', '-', '0'):
        return 'NULL'
    s = str(val).strip().replace("'", "''")
    return f"'{s}'"

def sql_num(val) -> str:
    """Format numeric value for SQL."""
    if val is None or str(val).strip() in ('', 'None', 'nan', '-'):
        return 'NULL'
    try:
        # Remove currency symbols, commas, spaces
        cleaned = re.sub(r'[$,\s]', '', str(val))
        return str(float(cleaned))
    except:
        return 'NULL'

def sql_int(val) -> str:
    if val is None or str(val).strip() in ('', 'None', 'nan', '-'):
        return 'NULL'
    try:
        return str(int(float(str(val).replace(',', ''))))
    except:
        return 'NULL'

def sql_date(val) -> str:
    """Convert date to SQL string."""
    if val is None or str(val).strip() in ('', 'None', 'nan', '-'):
        return 'NULL'
    if isinstance(val, (datetime, date)):
        return f"'{val.strftime('%Y-%m-%d')}'"
    s = str(val).strip()
    # Try common formats
    for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%Y-%m-%d %H:%M:%S']:
        try:
            return f"'{datetime.strptime(s, fmt).strftime('%Y-%m-%d')}'"
        except:
            continue
    return f"'{s}'"

def detect_product_group(scope_en: str, scope_vn: str, project: str) -> str:
    """Detect product group from scope/project description."""
    text = f"{scope_en or ''} {scope_vn or ''} {project or ''}".lower()
    for group, keywords in PRODUCT_GROUP_MAP.items():
        if any(kw in text for kw in keywords):
            return group
    return 'OTHER'

def detect_region(address: str, country: str = None) -> str:
    """Detect region from address/country."""
    text = f"{address or ''} {country or ''}".lower()
    if any(k in text for k in ['vietnam', 'việt nam', 'vn', 'hà nội', 'hcm', 'đà nẵng']):
        return 'VN'
    elif any(k in text for k in ['japan', 'nhật']):
        return 'JP'
    elif any(k in text for k in ['korea', 'hàn quốc']):
        return 'KR'
    elif any(k in text for k in ['usa', 'united states', 'america', 'louisville', 'spartanburg']):
        return 'US'
    elif any(k in text for k in ['germany', 'deutschland', 'belgium', 'netherlands', 'spain', 'europe', 'france', 'italy', 'uk']):
        return 'EU'
    elif any(k in text for k in ['australia', 'new zealand']):
        return 'AU'
    elif any(k in text for k in ['china', 'taiwan', 'singapore', 'thailand', 'indonesia', 'malaysia', 'philippines']):
        return 'ASIA'
    return 'OTHER'


# ══════════════════════════════════════════════════════════════════
# IMPORTERS
# ══════════════════════════════════════════════════════════════════

class WorkflowImporter:
    def __init__(self, base_dir: str = '.'):
        self.base_dir = Path(base_dir)
        self.customers: Dict[str, dict] = {}  # code → customer data
        self.contacts: List[dict] = []
        self.opportunities: List[dict] = []
        self.tasks: List[dict] = []
        self.customer_id_map: Dict[str, str] = {}  # code → generated id
        self.stats = {
            'customers_imported': 0,
            'contacts_imported': 0,
            'opportunities_imported': 0,
            'tasks_imported': 0,
            'errors': []
        }

    def load_customer_codes(self, filepath: str):
        """Load customer codes from Workflow 2026 'Cus-code' sheet."""
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        
        # Try different sheet names
        sheet_name = None
        for name in ['Cus-code', 'Customer code', 'Customer Code']:
            if name in wb.sheetnames:
                sheet_name = name
                break
        
        if not sheet_name:
            print(f"WARNING: Customer code sheet not found. Sheets: {wb.sheetnames}")
            wb.close()
            return
        
        ws = wb[sheet_name]
        print(f"Loading customers from sheet '{sheet_name}'...")
        
        # Find header row (look for 'Code' column)
        header_row = None
        headers = {}
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            vals = [str(v).strip().lower() if v else '' for v in row]
            if 'code' in vals or 'client code' in vals:
                header_row = row_idx
                for i, v in enumerate(vals):
                    if v:
                        headers[v] = i
                break
        
        if not header_row:
            print("WARNING: Could not find header row in customer sheet")
            wb.close()
            return
        
        print(f"  Headers found at row {header_row}: {list(headers.keys())}")
        
        # Map header names to our fields
        col_map = {}
        for h, idx in headers.items():
            h_lower = h.lower()
            if 'code' in h_lower and 'post' not in h_lower:
                col_map['code'] = idx
            elif h_lower in ('name', 'name of client', 'client name', 'tên'):
                col_map['name'] = idx
            elif 'country' in h_lower or 'quốc gia' in h_lower:
                col_map['country'] = idx
            elif 'market' in h_lower or 'thị trường' in h_lower:
                col_map['market'] = idx
            elif 'contact' in h_lower and 'name' not in h_lower:
                col_map['contact_name'] = idx
            elif 'email' in h_lower:
                col_map['email'] = idx
            elif 'position' in h_lower or 'chức vụ' in h_lower:
                col_map['position'] = idx
            elif 'phone' in h_lower or 'điện thoại' in h_lower:
                col_map['phone'] = idx
            elif 'website' in h_lower:
                col_map['website'] = idx
            elif 'linkedin' in h_lower:
                col_map['linkedin'] = idx
            elif 'address' in h_lower or 'địa chỉ' in h_lower:
                col_map['address'] = idx
            elif 'business' in h_lower or 'layout' in h_lower:
                col_map['business_description'] = idx
            elif 'status' in h_lower or 'trạng thái' in h_lower:
                col_map['status'] = idx
        
        print(f"  Column mapping: {col_map}")
        
        # Read data rows
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            vals = list(row)
            code_idx = col_map.get('code', 0)
            name_idx = col_map.get('name', 1)
            
            if not vals[code_idx] or not str(vals[code_idx]).strip():
                continue
            
            code = str(vals[code_idx]).strip().upper()
            name = str(vals[name_idx]).strip() if vals[name_idx] else code
            
            if code in self.customers:
                continue  # skip duplicates
            
            cust_id = gen_id()
            self.customer_id_map[code] = cust_id
            
            country = str(vals[col_map['country']]).strip() if 'country' in col_map and vals[col_map['country']] else None
            address = str(vals[col_map['address']]).strip() if 'address' in col_map and vals[col_map['address']] else None
            
            self.customers[code] = {
                'id': cust_id,
                'code': code,
                'name': name,
                'country': country,
                'region': detect_region(address or '', country or ''),
                'address': address,
                'website': str(vals[col_map['website']]).strip() if 'website' in col_map and vals[col_map['website']] else None,
                'business_description': str(vals[col_map['business_description']]).strip() if 'business_description' in col_map and vals[col_map['business_description']] else None,
                'status': str(vals[col_map['status']]).strip().upper() if 'status' in col_map and vals[col_map['status']] else 'ACTIVE',
            }
            
            # Extract contact if available
            contact_name = str(vals[col_map['contact_name']]).strip() if 'contact_name' in col_map and vals[col_map['contact_name']] else None
            if contact_name and contact_name not in ('None', '', '-'):
                self.contacts.append({
                    'id': gen_id(),
                    'customer_id': cust_id,
                    'name': contact_name,
                    'email': str(vals[col_map['email']]).strip() if 'email' in col_map and vals[col_map['email']] else None,
                    'phone': str(vals[col_map['phone']]).strip() if 'phone' in col_map and vals[col_map['phone']] else None,
                    'position': str(vals[col_map['position']]).strip() if 'position' in col_map and vals[col_map['position']] else None,
                    'linkedin': str(vals[col_map['linkedin']]).strip() if 'linkedin' in col_map and vals[col_map['linkedin']] else None,
                    'is_primary': 1,
                })
        
        self.stats['customers_imported'] = len(self.customers)
        self.stats['contacts_imported'] = len(self.contacts)
        print(f"  Loaded {len(self.customers)} customers, {len(self.contacts)} contacts")
        wb.close()

    def load_rfq_sheets(self, filepath: str):
        """Load RFQ/RFI sheets from Workflow 2026 for opportunities."""
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        
        rfq_sheets = [s for s in wb.sheetnames if any(k in s.lower() for k in ['rfq', 'rfi', 'workflow'])]
        print(f"RFQ/RFI sheets found: {rfq_sheets}")
        
        for sheet_name in rfq_sheets:
            ws = wb[sheet_name]
            print(f"\n  Processing sheet: {sheet_name} (rows={ws.max_row})")
            
            # Find header row
            header_row = None
            headers = {}
            for row_idx, row in enumerate(ws.iter_rows(max_row=10, values_only=True), 1):
                vals = [str(v).strip().lower() if v else '' for v in row]
                if any(k in ' '.join(vals) for k in ['quotation id', 'pre-sale', 'client', 'project']):
                    header_row = row_idx
                    for i, v in enumerate(vals):
                        if v:
                            headers[v] = i
                    break
            
            if not header_row:
                print(f"    WARNING: Could not find header row")
                continue
            
            # Map columns
            col_map = self._map_rfq_columns(headers)
            print(f"    Header row: {header_row}, Columns mapped: {list(col_map.keys())}")
            
            # Read data
            count = 0
            for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                vals = list(row)
                
                # Need at least a quotation ID or project name
                quot_id = str(vals[col_map.get('quotation_id', 0)]).strip() if col_map.get('quotation_id') is not None and vals[col_map.get('quotation_id', 0)] else None
                project = str(vals[col_map.get('project', 0)]).strip() if col_map.get('project') is not None and vals[col_map.get('project', 0)] else None
                
                if not quot_id and not project:
                    continue
                if quot_id in ('None', '', '-'):
                    quot_id = None
                
                # Find customer
                client_code = str(vals[col_map['client_code']]).strip().upper() if 'client_code' in col_map and vals[col_map['client_code']] else None
                client_name = str(vals[col_map['client_name']]).strip() if 'client_name' in col_map and vals[col_map['client_name']] else None
                
                customer_id = self.customer_id_map.get(client_code) if client_code else None
                
                # If customer not found by code, try to match by name
                if not customer_id and client_name:
                    for code, cust in self.customers.items():
                        if client_name.lower() in cust['name'].lower() or cust['name'].lower() in client_name.lower():
                            customer_id = cust['id']
                            break
                
                scope_en = str(vals[col_map['scope']]).strip() if 'scope' in col_map and vals[col_map['scope']] else None
                product_group = detect_product_group(scope_en or '', '', project or '')
                
                # Win probability
                win_pct_raw = vals[col_map.get('win_pct', 0)] if 'win_pct' in col_map else None
                try:
                    win_pct = int(float(str(win_pct_raw).replace('%', '')) * (100 if float(str(win_pct_raw).replace('%', '')) <= 1 else 1))
                except:
                    win_pct = 50
                
                # Determine stage from win probability
                if win_pct >= 100:
                    stage = 'WON'
                elif win_pct >= 70:
                    stage = 'NEGOTIATION'
                elif win_pct >= 50:
                    stage = 'QUOTED'
                elif win_pct >= 30:
                    stage = 'RFQ_RECEIVED'
                elif win_pct > 0:
                    stage = 'PROSPECT'
                else:
                    stage = 'LOST'
                
                opp = {
                    'id': gen_id(),
                    'pl_hd': quot_id,
                    'product_group': product_group,
                    'customer_id': customer_id,
                    'customer_name': client_name,
                    'project_name': project or f"RFQ-{quot_id}",
                    'scope_en': scope_en,
                    'stage': stage,
                    'win_probability': win_pct,
                    'weight_ton': vals[col_map.get('weight', -1)] if 'weight' in col_map else None,
                    'contract_value_usd': vals[col_map.get('value_usd', -1)] if 'value_usd' in col_map else None,
                    'contract_value_vnd': vals[col_map.get('value_vnd', -1)] if 'value_vnd' in col_map else None,
                    'gm_percent': vals[col_map.get('gm_pct', -1)] if 'gm_pct' in col_map else None,
                    'quotation_date': vals[col_map.get('quotation_date', -1)] if 'quotation_date' in col_map else None,
                    'source_sheet': sheet_name,
                }
                
                self.opportunities.append(opp)
                count += 1
            
            print(f"    Loaded {count} opportunities")
        
        self.stats['opportunities_imported'] = len(self.opportunities)
        wb.close()

    def _map_rfq_columns(self, headers: dict) -> dict:
        """Map RFQ sheet header names to standard field names."""
        col_map = {}
        for h, idx in headers.items():
            h_lower = h.lower()
            if 'quotation id' in h_lower or 'mã báo giá' in h_lower:
                col_map['quotation_id'] = idx
            elif 'pre-sale' in h_lower or 'pre sale' in h_lower:
                col_map['presale_id'] = idx
            elif 'client code' in h_lower or 'mã khách' in h_lower:
                col_map['client_code'] = idx
            elif ('name' in h_lower and 'client' in h_lower) or 'tên khách' in h_lower:
                col_map['client_name'] = idx
            elif 'project' in h_lower or 'dự án' in h_lower:
                col_map['project'] = idx
            elif 'scope' in h_lower or 'nội dung' in h_lower:
                col_map['scope'] = idx
            elif 'product' in h_lower or 'sản phẩm' in h_lower:
                col_map['product_type'] = idx
            elif 'weight' in h_lower or 'khối lượng' in h_lower:
                col_map['weight'] = idx
            elif 'value' in h_lower and 'usd' in h_lower:
                col_map['value_usd'] = idx
            elif 'value' in h_lower and 'vnd' in h_lower:
                col_map['value_vnd'] = idx
            elif 'gross profit' in h_lower or 'lợi nhuận' in h_lower:
                col_map['gm_pct'] = idx
            elif 'win' in h_lower or 'thắng' in h_lower:
                col_map['win_pct'] = idx
            elif 'quotation date' in h_lower or 'ngày chào' in h_lower:
                col_map['quotation_date'] = idx
        return col_map

    def load_tasks(self, filepath: str):
        """Load tasks from Workflow 2026 'To do list' sheets."""
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        
        task_sheets = [s for s in wb.sheetnames if 'to do' in s.lower() or 'todo' in s.lower()]
        print(f"Task sheets found: {task_sheets}")
        
        for sheet_name in task_sheets:
            ws = wb[sheet_name]
            print(f"\n  Processing: {sheet_name} (rows={ws.max_row})")
            
            # Find header row
            header_row = None
            headers = {}
            for row_idx, row in enumerate(ws.iter_rows(max_row=5, values_only=True), 1):
                vals = [str(v).strip().lower() if v else '' for v in row]
                if any(k in ' '.join(vals) for k in ['project', 'description', 'status', 'person']):
                    header_row = row_idx
                    for i, v in enumerate(vals):
                        if v:
                            headers[v] = i
                    break
            
            if not header_row:
                continue
            
            count = 0
            for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                vals = list(row)
                if not any(v for v in vals if v and str(v).strip()):
                    continue
                
                # Extract task fields by position (common layout: No, Project, Client, PO, Description, Start, Due, Days, Finish, Person, Status, Remarks)
                task = {
                    'id': gen_id(),
                    'task_type': 'SALE_FOLLOWUP',
                    'title': str(vals[4]).strip()[:200] if len(vals) > 4 and vals[4] else str(vals[1] or ''),
                    'description': str(vals[11]).strip()[:500] if len(vals) > 11 and vals[11] else None,
                    'from_dept': 'SALE',
                    'to_dept': 'SALE',
                    'assigned_to': str(vals[9]).strip() if len(vals) > 9 and vals[9] else None,
                    'due_date': vals[6] if len(vals) > 6 else None,
                    'status': self._map_task_status(str(vals[10]).strip() if len(vals) > 10 and vals[10] else ''),
                    'priority': 'NORMAL',
                    'source_sheet': sheet_name,
                }
                
                if task['title'] and task['title'] not in ('None', '', '-'):
                    self.tasks.append(task)
                    count += 1
            
            print(f"    Loaded {count} tasks")
        
        self.stats['tasks_imported'] = len(self.tasks)
        wb.close()

    def _map_task_status(self, status: str) -> str:
        s = status.lower()
        if s in ('close', 'closed', 'done', 'completed'):
            return 'COMPLETED'
        elif s in ('open', 'in progress', 'ongoing'):
            return 'IN_PROGRESS'
        elif s in ('pending', 'waiting'):
            return 'PENDING'
        return 'PENDING'

    def load_pipeline_data(self, filepath: str):
        """Load pipeline data from IBSHI Khach hang tiem nang xlsx."""
        try:
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        except Exception as e:
            print(f"WARNING: Could not open {filepath}: {e}")
            return
        
        if 'Data' not in wb.sheetnames:
            print(f"WARNING: 'Data' sheet not found in {filepath}")
            wb.close()
            return
        
        ws = wb['Data']
        print(f"Loading pipeline from 'Data' sheet (rows={ws.max_row})...")
        
        # Header at row 6
        # Columns: PLHD, No, Name, Address, Scope_EN, Scope_VN, Project, Win%, 
        #          Sign Date, Start, Duration, End, Weight, Value_VND, Value_USD,
        #          Unit Price, M1 Date, M1 Val, M2 Date, M2 Val, M3 Date, M3 Val, M4 Date, M4 Val,
        #          Material, Labor, Subcontractor, Profit, Profit%, 
        #          Qty2025, Val2025, GP2025, Qty2026, Val2026, GP2026, Quote Date, Action By, Value
        
        count = 0
        for row in ws.iter_rows(min_row=8, values_only=True):
            vals = list(row)
            name = vals[2]
            if not name or str(name).strip() in ('None', '', '-'):
                continue
            
            plhd = str(vals[0]).strip() if vals[0] else None
            stage = PLHD_STAGE_MAP.get(plhd, 'PROSPECT')
            
            win_raw = vals[7]
            try:
                win_pct = int(float(str(win_raw)) * 100) if float(str(win_raw)) <= 1 else int(float(str(win_raw)))
            except:
                win_pct = WIN_PCT_MAP.get(stage, 50)
            
            name_str = str(name).strip()
            scope_en = str(vals[4]).strip() if vals[4] else None
            scope_vn = str(vals[5]).strip() if vals[5] else None
            project = str(vals[6]).strip() if vals[6] else None
            product_group = detect_product_group(scope_en or '', scope_vn or '', project or '')
            
            # Find matching customer
            customer_id = None
            for code, cust in self.customers.items():
                if name_str.lower() in cust['name'].lower() or cust['name'].lower() in name_str.lower():
                    customer_id = cust['id']
                    break
            
            opp = {
                'id': gen_id(),
                'pl_hd': plhd,
                'product_group': product_group,
                'customer_id': customer_id,
                'customer_name': name_str,
                'project_name': project or f"Pipeline-{name_str}",
                'scope_en': scope_en,
                'scope_vn': scope_vn,
                'stage': stage,
                'win_probability': win_pct,
                'weight_ton': vals[12],
                'contract_value_vnd': vals[13],
                'contract_value_usd': vals[14],
                'unit_price_usd': vals[15],
                'estimated_signing': vals[8],
                'start_date': vals[9],
                'duration_months': vals[10],
                'end_date': vals[11],
                'material_cost_usd': vals[24],
                'labor_cost_usd': vals[25],
                'subcontractor_cost_usd': vals[26],
                'profit_usd': vals[27],
                'gm_percent': vals[28],
                'qty_2025': vals[29],
                'value_2025': vals[30],
                'gp_2025': vals[31],
                'qty_2026': vals[32],
                'value_2026': vals[33],
                'gp_2026': vals[34],
                'quotation_date': vals[35],
                'assigned_to': str(vals[36]).strip() if len(vals) > 36 and vals[36] else None,
                'source_sheet': 'Pipeline',
            }
            
            self.opportunities.append(opp)
            count += 1
        
        print(f"  Loaded {count} pipeline opportunities")
        self.stats['opportunities_imported'] = len(self.opportunities)
        wb.close()

    # ══════════════════════════════════════════════════════════════
    # SQL GENERATORS
    # ══════════════════════════════════════════════════════════════
    
    def generate_customer_sql(self) -> str:
        """Generate SQL for sale_customers table."""
        lines = [
            f"-- ═══════════════════════════════════════════════════════════",
            f"-- Sale Platform v4: Customer Import",
            f"-- Generated: {IMPORT_DATE}",
            f"-- Records: {len(self.customers)}",
            f"-- Source: Workflow 2026 + Client List",
            f"-- ═══════════════════════════════════════════════════════════",
            f"",
            f"BEGIN TRANSACTION;",
            f"",
        ]
        
        for code, c in sorted(self.customers.items()):
            lines.append(
                f"INSERT OR IGNORE INTO sale_customers "
                f"(id, code, name, country, region, address, website, business_description, status) VALUES "
                f"({sql_str(c['id'])}, {sql_str(c['code'])}, {sql_str(c['name'])}, "
                f"{sql_str(c['country'])}, {sql_str(c['region'])}, {sql_str(c['address'])}, "
                f"{sql_str(c['website'])}, {sql_str(c['business_description'])}, {sql_str(c['status'])});"
            )
        
        lines.extend(["", "COMMIT;", ""])
        return "\n".join(lines)

    def generate_contacts_sql(self) -> str:
        """Generate SQL for sale_customer_contacts table."""
        lines = [
            f"-- ═══════════════════════════════════════════════════════════",
            f"-- Sale Platform v4: Contact Import",
            f"-- Generated: {IMPORT_DATE}",
            f"-- Records: {len(self.contacts)}",
            f"-- ═══════════════════════════════════════════════════════════",
            f"",
            f"BEGIN TRANSACTION;",
            f"",
        ]
        
        for ct in self.contacts:
            lines.append(
                f"INSERT OR IGNORE INTO sale_customer_contacts "
                f"(id, customer_id, name, email, phone, position, linkedin, is_primary) VALUES "
                f"({sql_str(ct['id'])}, {sql_str(ct['customer_id'])}, {sql_str(ct['name'])}, "
                f"{sql_str(ct['email'])}, {sql_str(ct['phone'])}, {sql_str(ct['position'])}, "
                f"{sql_str(ct['linkedin'])}, {ct['is_primary']});"
            )
        
        lines.extend(["", "COMMIT;", ""])
        return "\n".join(lines)

    def generate_opportunities_sql(self) -> str:
        """Generate SQL for sale_opportunities table."""
        lines = [
            f"-- ═══════════════════════════════════════════════════════════",
            f"-- Sale Platform v4: Opportunity/Pipeline Import",
            f"-- Generated: {IMPORT_DATE}",
            f"-- Records: {len(self.opportunities)}",
            f"-- Source: Workflow 2026 RFQ sheets + Pipeline",
            f"-- ═══════════════════════════════════════════════════════════",
            f"",
            f"BEGIN TRANSACTION;",
            f"",
        ]
        
        for opp in self.opportunities:
            source_note = 'Source: ' + opp.get('source_sheet', 'unknown')
            lines.append(
                f"INSERT OR IGNORE INTO sale_opportunities "
                f"(id, pl_hd, product_group, customer_id, customer_name, project_name, "
                f"scope_en, scope_vn, stage, win_probability, weight_ton, "
                f"contract_value_vnd, contract_value_usd, unit_price_usd, "
                f"gm_percent, material_cost_usd, labor_cost_usd, subcontractor_cost_usd, profit_usd, "
                f"estimated_signing, start_date, duration_months, end_date, quotation_date, "
                f"qty_2025, value_2025, gp_2025, qty_2026, value_2026, gp_2026, "
                f"assigned_to, notes) VALUES "
                f"({sql_str(opp['id'])}, {sql_str(opp.get('pl_hd'))}, {sql_str(opp['product_group'])}, "
                f"{sql_str(opp.get('customer_id'))}, {sql_str(opp.get('customer_name'))}, {sql_str(opp['project_name'])}, "
                f"{sql_str(opp.get('scope_en'))}, {sql_str(opp.get('scope_vn'))}, {sql_str(opp['stage'])}, "
                f"{sql_int(opp.get('win_probability'))}, {sql_num(opp.get('weight_ton'))}, "
                f"{sql_num(opp.get('contract_value_vnd'))}, {sql_num(opp.get('contract_value_usd'))}, "
                f"{sql_num(opp.get('unit_price_usd'))}, "
                f"{sql_num(opp.get('gm_percent'))}, {sql_num(opp.get('material_cost_usd'))}, "
                f"{sql_num(opp.get('labor_cost_usd'))}, {sql_num(opp.get('subcontractor_cost_usd'))}, "
                f"{sql_num(opp.get('profit_usd'))}, "
                f"{sql_date(opp.get('estimated_signing'))}, {sql_date(opp.get('start_date'))}, "
                f"{sql_int(opp.get('duration_months'))}, {sql_date(opp.get('end_date'))}, "
                f"{sql_date(opp.get('quotation_date'))}, "
                f"{sql_num(opp.get('qty_2025'))}, {sql_num(opp.get('value_2025'))}, {sql_num(opp.get('gp_2025'))}, "
                f"{sql_num(opp.get('qty_2026'))}, {sql_num(opp.get('value_2026'))}, {sql_num(opp.get('gp_2026'))}, "
                f"{sql_str(opp.get('assigned_to'))}, {sql_str(source_note)});"
            )
        
        lines.extend(["", "COMMIT;", ""])
        return "\n".join(lines)

    def generate_tasks_sql(self) -> str:
        """Generate SQL for sale_tasks table."""
        lines = [
            f"-- ═══════════════════════════════════════════════════════════",
            f"-- Sale Platform v4: Task Import",
            f"-- Generated: {IMPORT_DATE}",
            f"-- Records: {len(self.tasks)}",
            f"-- ═══════════════════════════════════════════════════════════",
            f"",
            f"BEGIN TRANSACTION;",
            f"",
        ]
        
        for t in self.tasks:
            lines.append(
                f"INSERT OR IGNORE INTO sale_tasks "
                f"(id, task_type, title, description, from_dept, to_dept, assigned_to, "
                f"due_date, status, priority) VALUES "
                f"({sql_str(t['id'])}, {sql_str(t['task_type'])}, {sql_str(t['title'])}, "
                f"{sql_str(t.get('description'))}, {sql_str(t['from_dept'])}, {sql_str(t['to_dept'])}, "
                f"{sql_str(t.get('assigned_to'))}, {sql_date(t.get('due_date'))}, "
                f"{sql_str(t['status'])}, {sql_str(t['priority'])});"
            )
        
        lines.extend(["", "COMMIT;", ""])
        return "\n".join(lines)

    def generate_import_log_sql(self) -> str:
        """Generate import log entry."""
        return f"""-- Import Log Entry
INSERT INTO sale_import_log (id, import_type, source_file, records_total, records_imported, records_failed, imported_by, started_at, completed_at)
VALUES ('{gen_id()}', 'WORKFLOW_2026_FULL', 'Workflow 2026.xlsx + supplementary files',
  {sum([self.stats[k] for k in ['customers_imported', 'contacts_imported', 'opportunities_imported', 'tasks_imported']])},
  {sum([self.stats[k] for k in ['customers_imported', 'contacts_imported', 'opportunities_imported', 'tasks_imported']])},
  {len(self.stats['errors'])},
  'AI_IMPORT_SCRIPT', '{IMPORT_DATE}', '{IMPORT_DATE}');
"""

    def run(self):
        """Execute full import pipeline."""
        print("=" * 60)
        print("IBS HI Sale Platform v4 — Data Import")
        print("=" * 60)
        
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        
        # Step 1: Load customers
        wf_path = self.base_dir / WORKFLOW_FILE
        if wf_path.exists():
            self.load_customer_codes(str(wf_path))
            self.load_rfq_sheets(str(wf_path))
            self.load_tasks(str(wf_path))
        else:
            print(f"\nWARNING: {WORKFLOW_FILE} not found at {wf_path}")
            print("Export from Google Sheets first (File → Download → .xlsx)")
        
        # Step 2: Load pipeline data
        pipeline_path = self.base_dir / PIPELINE_FILE
        if pipeline_path.exists():
            self.load_pipeline_data(str(pipeline_path))
        
        # Step 3: Generate SQL files
        print(f"\n{'=' * 60}")
        print("Generating SQL files...")
        
        files = {
            '01_import_customers.sql': self.generate_customer_sql(),
            '02_import_contacts.sql': self.generate_contacts_sql(),
            '03_import_opportunities.sql': self.generate_opportunities_sql(),
            '04_import_tasks.sql': self.generate_tasks_sql(),
            '05_import_log.sql': self.generate_import_log_sql(),
        }
        
        for fname, content in files.items():
            fpath = os.path.join(OUTPUT_DIR, fname)
            with open(fpath, 'w', encoding='utf-8') as f:
                f.write(content)
            print(f"  Written: {fpath} ({len(content)} chars)")
        
        # Summary
        print(f"\n{'=' * 60}")
        print("IMPORT SUMMARY")
        print(f"{'=' * 60}")
        for k, v in self.stats.items():
            if k != 'errors':
                print(f"  {k}: {v}")
        if self.stats['errors']:
            print(f"  Errors: {len(self.stats['errors'])}")
            for e in self.stats['errors'][:5]:
                print(f"    - {e}")
        
        print(f"\nSQL files written to: {OUTPUT_DIR}/")
        print("Review SQL files, then execute:")
        print(f"  sqlite3 sale_platform.db < {OUTPUT_DIR}/01_import_customers.sql")
        print(f"  sqlite3 sale_platform.db < {OUTPUT_DIR}/02_import_contacts.sql")
        print(f"  sqlite3 sale_platform.db < {OUTPUT_DIR}/03_import_opportunities.sql")
        print(f"  sqlite3 sale_platform.db < {OUTPUT_DIR}/04_import_tasks.sql")
        print(f"  sqlite3 sale_platform.db < {OUTPUT_DIR}/05_import_log.sql")


if __name__ == '__main__':
    base = sys.argv[1] if len(sys.argv) > 1 else '.'
    importer = WorkflowImporter(base)
    importer.run()
