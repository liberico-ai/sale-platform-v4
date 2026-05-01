#!/usr/bin/env python3
"""
IBS HI Sale Platform v2.1 - Database Import Pipeline
Reads real data from NAS/GDrive and generates ready-to-run SQL files.
Run from the Sale/ folder root: python3 sale_db_import.py

Generated: 2026-04-28
"""
import os, sys, re, csv, json, uuid, sqlite3
from datetime import datetime, timedelta
from pathlib import Path

# ─── Configuration ───
# Adjust these paths to match your environment
SALE_ROOT = os.environ.get('SALE_ROOT', '.')
OUTPUT_DIR = os.environ.get('OUTPUT_DIR', './sql_import')
DB_PATH = os.environ.get('DB_PATH', './sale_platform_v4/sale.db')

# Data source paths (relative to SALE_ROOT)
PATHS = {
    'client_list': 'BaoCaoSale/Client List (2025) (2).xlsx',
    'pipeline_xlsx': 'BaoCaoSale/20250429_IBSHI Khách hàng tiềm năng_rev01.xlsx',
    'pipeline_csv': 'BaoCaoSale/20250429_IBSHI Khách hàng tiềm năng_rev01.csv',
    'sale_plan': '2025/TaiSaleDeXuat/Sale Plan 25-26-27 (Giá khoán).xlsx',
    'workload': '2025/IBS Workload 2025 -26 - 27.xlsx',
    'product_portfolio': 'HI ProductPorfolio.xlsx',
    'quotations_dir': '03.BaoGia_Quotations',
    'contracts_dir': '05.01.HopDongKhongGia_WithoutpriceContracts',
    'mkt_dir': 'MKT',
    'reports_dir': '02.BaoCao_Reports',
    'schema_sql': 'sale_platform_v4/schema.sql',
    'schema_new_sql': 'sale_platform_v4/schema_28_tables.sql',
}

def uid():
    return str(uuid.uuid4())[:16].replace('-','')

def esc(s):
    """Escape string for SQL — handles newlines, tabs, single quotes"""
    if s is None: return 'NULL'
    s = str(s).strip()
    if not s or s == 'None' or s == 'NULL': return 'NULL'
    s = s.replace("'", "''")
    s = s.replace('\n', ' ').replace('\r', ' ').replace('\t', ' ')
    # Collapse multiple spaces
    while '  ' in s:
        s = s.replace('  ', ' ')
    return "'" + s.strip() + "'"

def num(v):
    """Convert to number or NULL"""
    if v is None: return 'NULL'
    try:
        s = str(v).replace(',', '').replace(' ', '').strip()
        if not s or s == 'None' or s == 'NULL': return 'NULL'
        return str(round(float(s), 2))
    except: return 'NULL'

def iso_date(v):
    """Convert various date formats to ISO"""
    if v is None: return 'NULL'
    s = str(v).strip()
    if not s or s == 'None' or s == 'NULL': return 'NULL'
    # Already ISO
    if re.match(r'^\d{4}-\d{2}-\d{2}', s):
        return esc(s[:10])
    # DD/MM/YYYY
    m = re.match(r'(\d{1,2})/(\d{1,2})/(\d{4})', s)
    if m: return esc(f"{m.group(3)}-{m.group(2).zfill(2)}-{m.group(1).zfill(2)}")
    # Try datetime parsing
    for fmt in ['%d-%m-%Y', '%m/%d/%Y', '%Y/%m/%d', '%b %d, %Y', '%d %b %Y']:
        try:
            d = datetime.strptime(s, fmt)
            return esc(d.strftime('%Y-%m-%d'))
        except: continue
    return 'NULL'

# ─── Product Group Detection ───
PRODUCT_KEYWORDS = {
    'HRSG': ['hrsg', 'heat recovery', 'boiler', 'steam generator', 'nồi hơi', 'bypass', 'exhaust gas', 'diverter'],
    'DIVERTER': ['diverter', 'damper', 'valve', 'van chuyển'],
    'SHIP': ['ship', 'vessel', 'marine', 'hull', 'barge', 'tàu', 'docking', 'self-unloader', 'sul '],
    'PV': ['pressure vessel', 'column', 'tank', 'reactor', 'bình áp', 'autoclave', 'pv '],
    'HANDLING': ['conveyor', 'hopper', 'feeder', 'chute', 'bunker', 'silo', 'handling', 'crusher'],
    'DUCT': ['duct', 'stack', 'chimney', 'flue', 'ống dẫn', 'pipe', 'piping'],
}

def detect_product_group(scope):
    if not scope: return 'OTHER'
    s = str(scope).lower()
    for group, keywords in PRODUCT_KEYWORDS.items():
        if any(k in s for k in keywords):
            return group
    # Also check for structure/steel
    if any(k in s for k in ['structure', 'steel', 'frame', 'platform', 'support', 'monorail']):
        return 'HANDLING'
    return 'OTHER'

# ─── Region Detection ───
REGION_MAP = {
    'vietnam': 'VIETNAM', 'viet nam': 'VIETNAM', 'vn': 'VIETNAM', 'hải phòng': 'VIETNAM',
    'hà nội': 'VIETNAM', 'tp.hcm': 'VIETNAM', 'đà nẵng': 'VIETNAM',
    'japan': 'JAPAN', 'nhật': 'JAPAN', 'tokyo': 'JAPAN', 'osaka': 'JAPAN',
    'korea': 'ASIA', 'china': 'ASIA', 'taiwan': 'ASIA', 'singapore': 'ASIA',
    'india': 'ASIA', 'thailand': 'ASIA', 'indonesia': 'ASIA', 'malaysia': 'ASIA',
    'germany': 'EU', 'netherlands': 'EU', 'belgium': 'EU', 'france': 'EU',
    'denmark': 'EU', 'finland': 'EU', 'spain': 'EU', 'italy': 'EU', 'europe': 'EU',
    'usa': 'USA', 'united states': 'USA', 'america': 'USA',
    'australia': 'OCEANIA', 'new zealand': 'OCEANIA',
}

def detect_region(address):
    if not address: return None
    a = str(address).lower()
    for key, region in REGION_MAP.items():
        if key in a: return region
    return 'OTHER'

# ═══════════════════════════════════════════════════════════════
# IMPORTER CLASS
# ═══════════════════════════════════════════════════════════════
class SalePlatformImporter:
    def __init__(self, sale_root, output_dir):
        self.root = Path(sale_root)
        self.out = Path(output_dir)
        self.out.mkdir(exist_ok=True)
        self.customers = {}  # code -> {id, name, ...}
        self.opportunities = {}  # ref -> id
        self.stats = {}
        self.now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
    def log(self, msg):
        print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}")

    # ─── BATCH 1: Seed + Customers ───
    def import_clients(self):
        """Import from Client List 2025 + Sale Account Assignment"""
        self.log("Importing Client List...")
        lines = []
        try:
            import openpyxl
            path = self.root / PATHS['client_list']
            wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
            ws = wb['Sheet1']
            
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i < 2: continue  # Skip header rows
                vals = list(row)
                if not vals[2]: continue  # Skip empty rows
                
                code = str(vals[1]).strip() if vals[1] else None
                name = str(vals[2]).strip()
                if not code or code == 'None': continue
                
                # Determine account manager
                hieu = 'x' if vals[3] and str(vals[3]).strip().lower() == 'x' else None
                paul = 'x' if vals[4] and str(vals[4]).strip().lower() == 'x' else None
                assigned = 'Hiệu' if hieu else ('Paul' if paul else None)
                
                cid = uid()
                self.customers[code] = {'id': cid, 'name': name, 'code': code, 'assigned': assigned}
                
                lines.append(
                    f"INSERT OR IGNORE INTO sale_customers (id, code, name, status, created_at, updated_at) "
                    f"VALUES ({esc(cid)}, {esc(code)}, {esc(name)}, 'ACTIVE', {esc(self.now)}, {esc(self.now)});"
                )
            wb.close()
        except Exception as e:
            self.log(f"  ERROR reading Client List: {e}")
        
        # Also try Sale Account Assignment for extended client data
        try:
            path = self.root / PATHS['sale_plan']
            wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
            ws = wb['Sale Account Asignment']
            
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i < 4: continue  # Headers at row 3
                vals = list(row)
                code = str(vals[2]).strip() if vals[2] else None
                name = str(vals[3]).strip() if vals[3] else None
                if not code or code == 'None' or code in self.customers: continue
                
                hieu = 'x' if vals[4] and str(vals[4]).strip().lower() == 'x' else None
                paul = 'x' if vals[5] and str(vals[5]).strip().lower() == 'x' else None
                assigned = 'Hiệu' if hieu else ('Paul' if paul else None)
                
                cid = uid()
                self.customers[code] = {'id': cid, 'name': name, 'code': code, 'assigned': assigned}
                lines.append(
                    f"INSERT OR IGNORE INTO sale_customers (id, code, name, status, created_at, updated_at) "
                    f"VALUES ({esc(cid)}, {esc(code)}, {esc(name)}, 'ACTIVE', {esc(self.now)}, {esc(self.now)});"
                )
            wb.close()
        except Exception as e:
            self.log(f"  Note: Sale Account sheet: {e}")
        
        self.stats['customers'] = len(self.customers)
        self.log(f"  → {len(self.customers)} customers loaded")
        self._write_sql('01_customers.sql', lines, 'sale_customers')
        return lines

    def import_pipeline(self):
        """Import pipeline from IBSHI Khach hang tiem nang (Data sheet)"""
        self.log("Importing Pipeline Opportunities...")
        lines = []
        
        try:
            import openpyxl
            path = self.root / PATHS['pipeline_xlsx']
            wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
            ws = wb['Data']
            
            # Real headers at row 5 (index 5)
            # PL HD, STT, Ten KH, Dia chi, Noi dung CV, Noi dung CV, Ten du an,
            # Ty le thang thau, Thoi gian ky HD, Start, Duration, End,
            # Khoi luong (Tan), Gia tri VND, Gia tri USD, Don gia,
            # Milestone-1 Date, M1 Value, M2 Date, M2 Value, M3 Date, M3 Value, M4 Date, M4 Value,
            # Vat tu, Nhan cong, Chi phi thau phu, Loi nhuan con lai, %,
            # KL 2025, GT 2025, GP 2025, KL 2026, GT 2026, GP 2026
            
            opp_count = 0
            section_plhd = None  # carry-forward PLHD from section context

            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i < 6: continue  # Skip to data rows (header at row 5)
                vals = list(row)

                # Skip section headers and empty rows
                raw_plhd = str(vals[0]).strip() if vals[0] else None
                stt = vals[1]
                customer_name = str(vals[2]).strip() if vals[2] else None

                # Section header rows (STT = A, B, C) — track section context
                if stt and str(stt).strip() in ['A', 'B', 'C', 'D', 'E', 'NULL']:
                    sect = str(stt).strip()
                    if sect == 'A': section_plhd = 'I'    # WON/NEGOTIATION contracts
                    elif sect == 'B': section_plhd = 'II'  # High potential
                    elif sect == 'C': section_plhd = 'V'   # Other participating projects
                    continue

                if not customer_name or customer_name == 'None': continue
                if not stt: continue

                # Use explicit PLHD if present, otherwise inherit from section
                pl_hd = raw_plhd if raw_plhd else section_plhd
                
                address = str(vals[3]).strip() if vals[3] else None
                scope_vn = str(vals[4]).strip() if vals[4] else None
                scope_en = str(vals[5]).strip() if vals[5] else scope_vn
                project_name = str(vals[6]).strip() if vals[6] else f"{customer_name} - {scope_en or 'Project'}"
                win_pct = vals[7]
                est_signing = vals[8]
                start_date = vals[9]
                duration = vals[10]
                end_date = vals[11]
                weight = vals[12]
                value_vnd = vals[13]
                value_usd = vals[14]
                unit_price = vals[15] if len(vals) > 15 else None
                
                # Cost structure (cols 24-28)
                material_cost = vals[24] if len(vals) > 24 else None
                labor_cost = vals[25] if len(vals) > 25 else None
                sub_cost = vals[26] if len(vals) > 26 else None
                profit = vals[27] if len(vals) > 27 else None
                gm_pct = vals[28] if len(vals) > 28 else None
                
                # 2025 values (cols 29-31)
                qty_2025 = vals[29] if len(vals) > 29 else None
                val_2025 = vals[30] if len(vals) > 30 else None
                gp_2025 = vals[31] if len(vals) > 31 else None
                
                # 2026 values (cols 32-34)
                qty_2026 = vals[32] if len(vals) > 32 else None
                val_2026 = vals[33] if len(vals) > 33 else None
                gp_2026 = vals[34] if len(vals) > 34 else None
                
                # Detect product group
                product_group = detect_product_group(f"{scope_en} {scope_vn} {project_name}")
                region = detect_region(address)
                
                # Map customer
                customer_id = 'NULL'
                for code, cust in self.customers.items():
                    cname = cust.get('name') or ''
                    if customer_name and cname and cname.lower() in customer_name.lower():
                        customer_id = esc(cust['id'])
                        break
                    if customer_name and cname and customer_name.lower() in cname.lower():
                        customer_id = esc(cust['id'])
                        break
                
                # If no match, create new customer
                if customer_id == 'NULL' and customer_name:
                    cid = uid()
                    code_gen = re.sub(r'[^A-Z]', '', customer_name.upper())[:6]
                    if code_gen in self.customers:
                        code_gen = code_gen + str(opp_count)
                    self.customers[code_gen] = {'id': cid, 'name': customer_name, 'code': code_gen, 'assigned': None}
                    customer_id = esc(cid)
                    lines.append(
                        f"INSERT OR IGNORE INTO sale_customers (id, code, name, country, region, status, created_at, updated_at) "
                        f"VALUES ({esc(cid)}, {esc(code_gen)}, {esc(customer_name)}, NULL, {esc(region)}, 'PROSPECT', {esc(self.now)}, {esc(self.now)});"
                    )
                
                # Determine stage from PLHD and win%
                stage = 'PROSPECT'
                if pl_hd:
                    pl = str(pl_hd).upper().strip()
                    if 'I' == pl or pl == 'PLHD I': stage = 'WON'
                    elif 'II' == pl or pl == 'PLHD II': stage = 'NEGOTIATION'
                    elif 'III' == pl: stage = 'WON'
                    elif 'IV' == pl: stage = 'QUOTED'
                    elif 'V' == pl: stage = 'PROSPECT'
                
                # Override with win% if available
                try:
                    wp = float(str(win_pct).replace('%', '').replace(',', '.'))
                    if wp >= 0.9 or wp >= 90: stage = 'WON'
                    elif wp >= 0.7 or wp >= 70: stage = 'NEGOTIATION'
                    elif wp >= 0.4 or wp >= 40: stage = 'QUOTED'
                    elif wp >= 0.1 or wp >= 10: stage = 'RFQ_RECEIVED'
                    win_prob = int(wp * 100) if wp <= 1 else int(wp)
                except:
                    win_prob = 50
                
                oid = uid()
                self.opportunities[f"OPP-{opp_count+1}"] = oid
                opp_count += 1
                
                lines.append(
                    f"INSERT OR IGNORE INTO sale_opportunities "
                    f"(id, pl_hd, product_group, customer_id, customer_name, project_name, "
                    f"scope_en, scope_vn, stage, win_probability, weight_ton, "
                    f"contract_value_vnd, contract_value_usd, unit_price_usd, "
                    f"gm_percent, material_cost_usd, labor_cost_usd, subcontractor_cost_usd, profit_usd, "
                    f"estimated_signing, start_date, duration_months, end_date, "
                    f"qty_2025, value_2025, gp_2025, qty_2026, value_2026, gp_2026, "
                    f"assigned_to, notes, created_at, updated_at) VALUES "
                    f"({esc(oid)}, {esc(pl_hd)}, {esc(product_group)}, {customer_id}, {esc(customer_name)}, "
                    f"{esc(project_name)}, {esc(scope_en)}, {esc(scope_vn)}, "
                    f"{esc(stage)}, {num(win_prob)}, {num(weight)}, "
                    f"{num(value_vnd)}, {num(value_usd)}, {num(unit_price)}, "
                    f"{num(gm_pct)}, {num(material_cost)}, {num(labor_cost)}, {num(sub_cost)}, {num(profit)}, "
                    f"{iso_date(est_signing)}, {iso_date(start_date)}, {num(duration)}, {iso_date(end_date)}, "
                    f"{num(qty_2025)}, {num(val_2025)}, {num(gp_2025)}, {num(qty_2026)}, {num(val_2026)}, {num(gp_2026)}, "
                    f"NULL, 'Source: IBSHI Khach hang tiem nang Data sheet', {esc(self.now)}, {esc(self.now)});"
                )
            
            wb.close()
        except Exception as e:
            self.log(f"  ERROR: {e}")
            import traceback; traceback.print_exc()
        
        self.stats['opportunities'] = opp_count
        self.log(f"  → {opp_count} opportunities imported")
        self._write_sql('02_opportunities.sql', lines, 'sale_opportunities')
        return lines

    def import_active_contracts(self):
        """Import from Sale Plan 'Cac HĐ và dự án tiềm năng' → additional opportunities with project codes"""
        self.log("Importing Active Contracts & Projects (Sale Plan)...")
        lines = []

        try:
            import openpyxl
            path = self.root / PATHS['sale_plan']
            wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
            ws = wb['Cac HĐ và dự án tiềm năng']

            # Headers at row 2: col1=ProjectCode, col2=Customer, col3=Address, col4=ScopeEN,
            # col5=ScopeVN, col6=ProjectName, col7=WinRate, col8=SigningDate, col9=StartDate,
            # col10=Duration, col11=EndDate, col12=Weight(ton), col13=ValueVND, col14=ValueUSD,
            # col15=UnitPrice, col16=MaterialCost, col17=LaborCost, col18=SubCost, col19=GP,
            # col20=Qty2025, col21=Val2025, col22=Qty2026, col23=Val2026, col24=QuotationNo, col25=Remarks
            # cols 26+ are payment milestones (date, amount pairs)
            count = 0
            section = None
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i < 5: continue  # Skip to data (rows 0-4 are headers)
                vals = list(row)

                proj_code = str(vals[1]).strip() if vals[1] else None
                customer_name = str(vals[2]).strip() if vals[2] else None

                # Section headers
                if proj_code in ['A', 'B', 'I', 'II', 'III']:
                    if proj_code == 'A': section = 'SIGNED'
                    elif proj_code == 'I': section = 'SIGNED_2024'
                    elif proj_code == 'II': section = 'SIGNED_2025'
                    elif proj_code == 'B': section = 'POTENTIAL'
                    continue

                if not customer_name or customer_name == 'None': continue
                if not proj_code or proj_code == 'None': continue
                # Skip if proj_code is not a number (project code like 076, 078...)
                try: int(proj_code)
                except: continue

                address = str(vals[3]).strip() if vals[3] else None
                scope_en = str(vals[4]).strip() if vals[4] else None
                scope_vn = str(vals[5]).strip() if vals[5] else scope_en
                project_name = str(vals[6]).strip() if vals[6] else f"{customer_name} - {scope_en or 'Project'}"
                win_pct = vals[7]
                est_signing = vals[8]
                start_date = vals[9]
                duration = vals[10]
                end_date = vals[11]
                weight = vals[12]
                value_vnd = vals[13]
                value_usd = vals[14]
                unit_price = vals[15] if len(vals) > 15 else None
                material_cost = vals[16] if len(vals) > 16 else None
                labor_cost = vals[17] if len(vals) > 17 else None
                sub_cost = vals[18] if len(vals) > 18 else None
                gp = vals[19] if len(vals) > 19 else None
                qty_2025 = vals[20] if len(vals) > 20 else None
                val_2025 = vals[21] if len(vals) > 21 else None
                qty_2026 = vals[22] if len(vals) > 22 else None
                val_2026 = vals[23] if len(vals) > 23 else None
                quotation_no = str(vals[24]).strip() if len(vals) > 24 and vals[24] else None
                remarks = str(vals[25]).strip() if len(vals) > 25 and vals[25] else None

                # Determine stage
                stage = 'PROSPECT'
                try:
                    wp = float(str(win_pct).replace('%', '').replace(',', '.'))
                    if wp >= 0.9 or wp >= 90: stage = 'WON'
                    elif wp >= 0.7 or wp >= 70: stage = 'NEGOTIATION'
                    elif wp >= 0.4 or wp >= 40: stage = 'QUOTED'
                    win_prob = int(wp * 100) if wp <= 1 else int(wp)
                except:
                    win_prob = 50
                    if section in ['SIGNED', 'SIGNED_2024', 'SIGNED_2025']: stage = 'WON'

                product_group = detect_product_group(f"{scope_en} {scope_vn} {project_name}")
                region = detect_region(address)

                # Map customer
                customer_id = 'NULL'
                for code, cust in self.customers.items():
                    cname = cust.get('name') or ''
                    if customer_name and cname and cname.lower() in customer_name.lower():
                        customer_id = esc(cust['id'])
                        break
                    if customer_name and cname and customer_name.lower() in cname.lower():
                        customer_id = esc(cust['id'])
                        break

                if customer_id == 'NULL' and customer_name:
                    cid = uid()
                    code_gen = re.sub(r'[^A-Z]', '', customer_name.upper())[:6]
                    if code_gen in self.customers:
                        code_gen = code_gen + str(count)
                    self.customers[code_gen] = {'id': cid, 'name': customer_name, 'code': code_gen, 'assigned': None}
                    customer_id = esc(cid)
                    lines.append(
                        f"INSERT OR IGNORE INTO sale_customers (id, code, name, country, region, status, created_at, updated_at) "
                        f"VALUES ({esc(cid)}, {esc(code_gen)}, {esc(customer_name)}, NULL, {esc(region)}, 'ACTIVE', {esc(self.now)}, {esc(self.now)});"
                    )

                oid = uid()
                count += 1
                proj_ref = f"25-{proj_code}"
                notes_parts = [f'Source: Sale Plan / Cac HD va du an tiem nang']
                if quotation_no: notes_parts.append(f'Quotation: {quotation_no}')
                if remarks: notes_parts.append(f'Remarks: {remarks}')
                if section: notes_parts.append(f'Section: {section}')

                lines.append(
                    f"INSERT OR IGNORE INTO sale_opportunities "
                    f"(id, pl_hd, product_group, customer_id, customer_name, project_name, "
                    f"scope_en, scope_vn, stage, win_probability, weight_ton, "
                    f"contract_value_vnd, contract_value_usd, unit_price_usd, "
                    f"material_cost_usd, labor_cost_usd, subcontractor_cost_usd, profit_usd, "
                    f"estimated_signing, start_date, duration_months, end_date, "
                    f"qty_2025, value_2025, gp_2025, qty_2026, value_2026, gp_2026, "
                    f"assigned_to, notes, created_at, updated_at) VALUES "
                    f"({esc(oid)}, {esc(proj_ref)}, {esc(product_group)}, {customer_id}, {esc(customer_name)}, "
                    f"{esc(project_name)}, {esc(scope_en)}, {esc(scope_vn)}, "
                    f"{esc(stage)}, {num(win_prob)}, {num(weight)}, "
                    f"{num(value_vnd)}, {num(value_usd)}, {num(unit_price)}, "
                    f"{num(material_cost)}, {num(labor_cost)}, {num(sub_cost)}, {num(gp)}, "
                    f"{iso_date(est_signing)}, {iso_date(start_date)}, {num(duration)}, {iso_date(end_date)}, "
                    f"{num(qty_2025)}, {num(val_2025)}, NULL, {num(qty_2026)}, {num(val_2026)}, NULL, "
                    f"NULL, {esc('; '.join(notes_parts))}, {esc(self.now)}, {esc(self.now)});"
                )

                # Also extract payment milestones (cols 26+ are date/amount pairs)
                for m_idx in range(5):  # up to 5 milestones
                    d_col = 26 + m_idx * 2
                    v_col = 27 + m_idx * 2
                    if len(vals) > v_col and vals[d_col] and vals[v_col]:
                        mid = uid()
                        lines.append(
                            f"INSERT OR IGNORE INTO sale_contract_milestones "
                            f"(id, opportunity_id, milestone_number, milestone_type, title, "
                            f"planned_date, invoice_amount_vnd, invoice_status, notes, created_at, updated_at) VALUES "
                            f"({esc(mid)}, {esc(oid)}, {m_idx+1}, 'PAYMENT', "
                            f"{esc(f'Payment Milestone {m_idx+1} - {proj_ref}')}, "
                            f"{iso_date(vals[d_col])}, {num(vals[v_col])}, 'NOT_INVOICED', "
                            f"{esc(f'Auto-imported from Sale Plan')}, {esc(self.now)}, {esc(self.now)});"
                        )

            wb.close()
        except Exception as e:
            self.log(f"  ERROR: {e}")
            import traceback; traceback.print_exc()

        self.stats['active_contracts'] = count
        self.log(f"  → {count} active contracts/projects imported")
        self._write_sql('02b_active_contracts.sql', lines, 'sale_opportunities + sale_contract_milestones')

    def import_invoice_status(self):
        """Import from Sale Plan Invoice Status → contract_milestones"""
        self.log("Importing Invoice Status → contract milestones...")
        lines = []
        
        try:
            import openpyxl
            path = self.root / PATHS['sale_plan']
            wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
            ws = wb['Invoice Status']
            
            # Headers at row 6-7 (merged)
            # STT, _, Project Number, Content, Payment %, Contract No, Contract Value, 
            # Payment terms, Invoice No, Value, Currency, Plan dates...
            
            count = 0
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i < 8: continue  # Skip headers
                vals = list(row)
                
                stt = vals[0]
                project_num = str(vals[2]).strip() if vals[2] else None
                content = str(vals[3]).strip() if vals[3] else None
                payment_pct = vals[4]
                contract_no = str(vals[5]).strip() if vals[5] else None
                contract_value = vals[6]
                payment_terms = vals[7]
                invoice_no = str(vals[8]).strip() if vals[8] else None
                inv_value = vals[9]
                currency = str(vals[10]).strip() if vals[10] else 'USD'
                
                if not project_num or project_num == 'None': continue
                if not stt: continue
                
                mid = uid()
                count += 1
                
                # Determine milestone type from payment %
                m_type = 'DELIVERY'
                if payment_pct:
                    pct_str = str(payment_pct).lower()
                    if 'advance' in pct_str or 'tạm ứng' in pct_str: m_type = 'ADVANCE'
                    elif 'final' in pct_str or 'cuối' in pct_str: m_type = 'FINAL'
                    elif 'retention' in pct_str or 'bảo lãnh' in pct_str: m_type = 'RETENTION'
                
                # Determine status
                status = 'NOT_INVOICED'
                if invoice_no and invoice_no != 'None':
                    status = 'INVOICED'
                
                lines.append(
                    f"INSERT OR IGNORE INTO sale_contract_milestones "
                    f"(id, opportunity_id, milestone_number, milestone_type, title, description, "
                    f"invoice_number, invoice_amount_usd, invoice_amount_vnd, invoice_status, "
                    f"notes, created_at, updated_at) VALUES "
                    f"({esc(mid)}, NULL, {count}, {esc(m_type)}, {esc(f'{project_num} - {content}')}, "
                    f"{esc(content)}, {esc(invoice_no)}, "
                    f"{'NULL' if currency != 'USD' else num(inv_value)}, "
                    f"{'NULL' if currency == 'USD' else num(inv_value)}, "
                    f"{esc(status)}, {esc(f'Contract: {contract_no}, Project: {project_num}')}, "
                    f"{esc(self.now)}, {esc(self.now)});"
                )
            
            wb.close()
        except Exception as e:
            self.log(f"  ERROR: {e}")
        
        self.stats['milestones'] = count
        self.log(f"  → {count} contract milestones imported")
        self._write_sql('03_contract_milestones.sql', lines, 'sale_contract_milestones')

    def import_thu_hoi_von(self):
        """Import Capital Recovery from Sale Plan Thu hoi von sheet"""
        self.log("Importing Capital Recovery (Thu hoi von)...")
        lines = []
        
        try:
            import openpyxl
            path = self.root / PATHS['sale_plan']
            wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
            ws = wb['Thu hoi von']
            
            count = 0
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i < 5: continue  # Headers
                vals = list(row)
                
                project_no = str(vals[0]).strip() if vals[0] else None
                stt = vals[1]
                project_name = str(vals[2]).strip() if vals[2] else None
                contract_usd = vals[3]
                contract_vnd = vals[4]
                
                if not project_name or project_name == 'None': continue
                if not stt: continue
                # Skip subtotal/section rows
                if any(k in str(project_name).lower() for k in ['hợp đồng', 'tổng', 'total']): continue
                
                sid = uid()
                count += 1
                
                # Collect quarterly recovery values
                q1_usd = vals[5] if len(vals) > 5 else None
                q1_vnd = vals[6] if len(vals) > 6 else None
                
                lines.append(
                    f"INSERT OR IGNORE INTO sale_settlements "
                    f"(id, opportunity_id, settlement_status, planned_value_usd, "
                    f"total_invoiced, total_received, notes, created_at, updated_at) VALUES "
                    f"({esc(sid)}, NULL, 'IN_PROGRESS', {num(contract_usd)}, "
                    f"NULL, {num(q1_usd)}, {esc(f'Project: {project_no} - {project_name}')}, "
                    f"{esc(self.now)}, {esc(self.now)});"
                )
            
            wb.close()
        except Exception as e:
            self.log(f"  ERROR: {e}")
        
        self.stats['settlements'] = count
        self.log(f"  → {count} settlement records imported")
        self._write_sql('04_settlements.sql', lines, 'sale_settlements')

    def import_vogt_pipeline(self):
        """Import VOGT Pipeline → additional opportunities"""
        self.log("Importing VOGT Pipeline...")
        lines = []
        
        try:
            import openpyxl
            path = self.root / PATHS['sale_plan']
            wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
            ws = wb['VOGTPipeline']
            
            count = 0
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i < 4: continue  # Headers at row 3
                vals = list(row)
                
                stt = vals[1]
                project = str(vals[2]).strip() if vals[2] else None
                vendor_name = str(vals[3]).strip() if vals[3] else None
                country = str(vals[4]).strip() if vals[4] else None
                units = vals[5]
                release_date = str(vals[6]).strip() if vals[6] else None
                design_type = str(vals[7]).strip() if vals[7] else None
                npp_mt = vals[8]
                npp_vendor = str(vals[9]).strip() if vals[9] else None
                
                if not project or project == 'None': continue
                if not stt: continue
                
                oid = uid()
                count += 1
                
                # VOGT = HRSG product
                lines.append(
                    f"INSERT OR IGNORE INTO sale_opportunities "
                    f"(id, pl_hd, product_group, customer_name, project_name, "
                    f"scope_en, stage, weight_ton, assigned_to, "
                    f"notes, created_at, updated_at) VALUES "
                    f"({esc(oid)}, 'VOGT', 'HRSG', 'VOGT Power International', "
                    f"{esc(f'VOGT - {project}')}, {esc(f'{design_type} - {vendor_name} ({country})')}, "
                    f"'PROSPECT', {num(npp_mt)}, 'Hiệu', "
                    f"{esc(f'VOGT Pipeline. Units: {units}. Release: {release_date}. NPP: {npp_vendor}')}, "
                    f"{esc(self.now)}, {esc(self.now)});"
                )
            
            wb.close()
        except Exception as e:
            self.log(f"  ERROR: {e}")
        
        self.stats['vogt'] = count
        self.log(f"  → {count} VOGT pipeline records imported")
        self._write_sql('05_vogt_pipeline.sql', lines, 'sale_opportunities (VOGT)')

    def import_quotation_files(self):
        """Scan quotation folders → sale_nas_file_links + sale_quotation_revisions"""
        self.log("Scanning quotation folders...")
        file_lines = []
        rev_lines = []
        
        qt_dir = self.root / PATHS['quotations_dir']
        if not qt_dir.exists():
            self.log(f"  Quotation dir not found: {qt_dir}")
            return
        
        file_count = 0
        rev_count = 0
        
        for month_dir in sorted(qt_dir.iterdir()):
            if not month_dir.is_dir() or month_dir.name.startswith('.'): continue
            
            for item in sorted(month_dir.iterdir()):
                if item.name.startswith('.'): continue
                
                if item.is_dir():
                    # This is a quotation subfolder like "25001 (250113) Quotation for..."
                    qt_name = item.name
                    # Extract quotation number
                    qt_match = re.match(r'(\d{5})', qt_name)
                    qt_ref = qt_match.group(1) if qt_match else qt_name[:10]
                    
                    # Extract project description
                    desc_match = re.search(r'Quotation for (.+?)(?:\s*[-_]\s*\w+)?$', qt_name)
                    desc = desc_match.group(1) if desc_match else qt_name
                    
                    # Create revision record
                    rid = uid()
                    rev_count += 1
                    rev_lines.append(
                        f"INSERT OR IGNORE INTO sale_quotation_revisions "
                        f"(id, opportunity_id, revision_number, revision_date, "
                        f"quotation_ref, scope_summary, nas_file_path, "
                        f"created_at, updated_at) VALUES "
                        f"({esc(rid)}, NULL, 1, {esc(self.now[:10])}, "
                        f"{esc(qt_ref)}, {esc(desc)}, "
                        f"{esc(f'//NAS/KINH DOANH-KTKH/03.BaoGia_Quotations/{month_dir.name}/{qt_name}')}, "
                        f"{esc(self.now)}, {esc(self.now)});"
                    )
                    
                    # Index files within subfolder
                    for f in item.iterdir():
                        if f.name.startswith('.'): continue
                        fid = uid()
                        file_count += 1
                        file_lines.append(
                            f"INSERT OR IGNORE INTO sale_nas_file_links "
                            f"(id, entity_type, entity_id, nas_path, file_name, file_type, description, created_at) VALUES "
                            f"({esc(fid)}, 'QUOTATION', {esc(qt_ref)}, "
                            f"{esc(f'//NAS/KINH DOANH-KTKH/03.BaoGia_Quotations/{month_dir.name}/{qt_name}/{f.name}')}, "
                            f"{esc(f.name)}, {esc(f.suffix.lstrip('.').upper())}, "
                            f"{esc(f'Quotation {qt_ref}')}, {esc(self.now)});"
                        )
                elif item.is_file():
                    # Direct file in month folder
                    fid = uid()
                    file_count += 1
                    file_lines.append(
                        f"INSERT OR IGNORE INTO sale_nas_file_links "
                        f"(id, entity_type, entity_id, nas_path, file_name, file_type, description, created_at) VALUES "
                        f"({esc(fid)}, 'QUOTATION_FILE', {esc(month_dir.name)}, "
                        f"{esc(f'//NAS/KINH DOANH-KTKH/03.BaoGia_Quotations/{month_dir.name}/{item.name}')}, "
                        f"{esc(item.name)}, {esc(item.suffix.lstrip('.').upper())}, "
                        f"{esc(f'Quotation folder {month_dir.name}')}, {esc(self.now)});"
                    )
        
        self.stats['nas_files'] = file_count
        self.stats['revisions'] = rev_count
        self.log(f"  → {file_count} NAS file links, {rev_count} quotation revisions")
        self._write_sql('06_nas_file_links.sql', file_lines, 'sale_nas_file_links')
        self._write_sql('07_quotation_revisions.sql', rev_lines, 'sale_quotation_revisions')

    def import_contracts(self):
        """Scan contract folders → sale_nas_file_links"""
        self.log("Scanning contract folders...")
        lines = []
        
        ct_dir = self.root / PATHS['contracts_dir']
        if not ct_dir.exists():
            self.log(f"  Contract dir not found")
            return
        
        count = 0
        for customer_dir in sorted(ct_dir.iterdir()):
            if not customer_dir.is_dir() or customer_dir.name.startswith('.'): continue
            customer_label = customer_dir.name  # e.g., "1. VOGT"
            
            for contract_dir in sorted(customer_dir.iterdir()):
                if not contract_dir.is_dir() or contract_dir.name.startswith('.'): continue
                contract_ref = contract_dir.name  # e.g., "25-VPI-078 (V17556-PT FREEPORT)"
                
                for f in contract_dir.iterdir():
                    if f.name.startswith('.'): continue
                    fid = uid()
                    count += 1
                    lines.append(
                        f"INSERT OR IGNORE INTO sale_nas_file_links "
                        f"(id, entity_type, entity_id, nas_path, file_name, file_type, description, created_at) VALUES "
                        f"({esc(fid)}, 'CONTRACT', {esc(contract_ref)}, "
                        f"{esc(f'//NAS/KINH DOANH-KTKH/05.HopDong_Contracts/{customer_label}/{contract_ref}/{f.name}')}, "
                        f"{esc(f.name)}, {esc(f.suffix.lstrip('.').upper())}, "
                        f"{esc(f'Contract {contract_ref} - {customer_label}')}, {esc(self.now)});"
                    )
        
        self.stats['contract_files'] = count
        self.log(f"  → {count} contract file links")
        self._write_sql('08_contract_files.sql', lines, 'sale_nas_file_links (contracts)')

    def import_mkt_content(self):
        """Scan MKT folder → sale_digital_content"""
        self.log("Scanning MKT folder...")
        lines = []
        
        mkt_dir = self.root / PATHS['mkt_dir']
        if not mkt_dir.exists():
            self.log(f"  MKT dir not found")
            return
        
        count = 0
        for f in sorted(mkt_dir.iterdir()):
            if f.name.startswith('.') or not f.is_file(): continue
            
            did = uid()
            count += 1
            
            # Detect content type
            ext = f.suffix.lower()
            if ext in ['.pdf', '.pptx']:
                content_type = 'PRESENTATION' if ext == '.pptx' else 'BROCHURE'
                if 'company profile' in f.name.lower(): content_type = 'COMPANY_PROFILE'
                elif 'corporate presentation' in f.name.lower(): content_type = 'PRESENTATION'
            elif ext in ['.jpg', '.png']:
                content_type = 'INFOGRAPHIC'
            elif ext in ['.docx']:
                content_type = 'BROCHURE'
            else:
                content_type = 'BROCHURE'
            
            file_size = f.stat().st_size / (1024*1024)  # MB
            
            lines.append(
                f"INSERT OR IGNORE INTO sale_digital_content "
                f"(id, content_type, title, file_format, file_size_mb, "
                f"nas_file_path, version, is_latest, is_active, created_at, updated_at) VALUES "
                f"({esc(did)}, {esc(content_type)}, {esc(f.stem)}, "
                f"{esc(ext.lstrip('.').upper())}, {num(file_size)}, "
                f"{esc(f'//NAS/KINH DOANH-KTKH/10.Marketing/{f.name}')}, "
                f"1, 1, 1, {esc(self.now)}, {esc(self.now)});"
            )
        
        self.stats['mkt'] = count
        self.log(f"  → {count} marketing content records")
        self._write_sql('09_digital_content.sql', lines, 'sale_digital_content')

    def import_product_portfolio(self):
        """Import from HI ProductPortfolio.xlsx → sale_product_opportunities"""
        self.log("Importing Product Portfolio...")
        lines = []

        try:
            import openpyxl
            path = self.root / PATHS['product_portfolio']
            wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
            ws = wb['Sheet1']

            # Structure: R0=title, R1=blank, R2=headers (No, _, Sectors, Name of product, Difficulty, Productivity, Notes, Market price)
            # Data from R3+: col[2]=sector group, col[3]=product name, col[4]=difficulty, col[5]=productivity
            count = 0
            current_sector = None
            current_group = None
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i < 3: continue  # Skip title + header rows
                vals = list(row)

                # Track sector (col 1-2) when present
                if len(vals) > 2 and vals[2]:
                    current_sector = str(vals[2]).strip()
                if len(vals) > 1 and vals[1]:
                    s = str(vals[1]).strip()
                    if s and s != 'None':
                        current_sector = s

                product = str(vals[3]).strip() if len(vals) > 3 and vals[3] else None
                if not product or product == 'None': continue

                difficulty = str(vals[4]).strip() if len(vals) > 4 and vals[4] else None
                productivity = str(vals[5]).strip() if len(vals) > 5 and vals[5] else None
                notes_col = str(vals[6]).strip() if len(vals) > 6 and vals[6] else None

                pid = uid()
                count += 1

                group = detect_product_group(f"{product} {current_sector or ''}")
                notes_parts = [f'Product: {product}']
                if current_sector: notes_parts.append(f'Sector: {current_sector}')
                if difficulty: notes_parts.append(f'Difficulty: {difficulty}')
                if productivity: notes_parts.append(f'Productivity: {productivity}')
                if notes_col: notes_parts.append(f'Notes: {notes_col}')

                lines.append(
                    f"INSERT OR IGNORE INTO sale_product_opportunities "
                    f"(id, product_category_id, capability_status, notes, created_at, updated_at) VALUES "
                    f"({esc(pid)}, NULL, 'FULL', {esc('; '.join(notes_parts))}, {esc(self.now)}, {esc(self.now)});"
                )
            
            wb.close()
        except Exception as e:
            self.log(f"  ERROR: {e}")
        
        self.stats['products'] = count
        self.log(f"  → {count} product capability records")
        self._write_sql('10_product_opportunities.sql', lines, 'sale_product_opportunities')

    def generate_import_log(self):
        """Generate import audit log"""
        self.log("Generating import log...")
        lines = []
        
        for source, count in self.stats.items():
            lid = uid()
            lines.append(
                f"INSERT INTO sale_import_log "
                f"(id, import_type, source_file, records_total, records_imported, records_failed, "
                f"imported_by, started_at, completed_at) VALUES "
                f"({esc(lid)}, {esc(source)}, {esc(f'sale_db_import.py - {source}')}, "
                f"{count}, {count}, 0, 'AI Chief of Staff', {esc(self.now)}, {esc(self.now)});"
            )
        
        self._write_sql('99_import_log.sql', lines, 'sale_import_log')

    def generate_master_script(self):
        """Generate a master SQL that runs all imports in order"""
        self.log("Generating master import script...")
        
        # List all SQL files in order
        sql_files = sorted(self.out.glob('*.sql'))
        
        master = [
            f"-- ═══════════════════════════════════════════════════════════════",
            f"-- IBS HI SALE PLATFORM - MASTER IMPORT SCRIPT",
            f"-- Generated: {self.now}",
            f"-- Run: sqlite3 sale.db < master_import.sql",
            f"-- ═══════════════════════════════════════════════════════════════",
            f"",
            f"-- First, create the 12 new tables if they don't exist",
            f"-- .read schema_28_tables.sql",
            f"",
            f"BEGIN TRANSACTION;",
            f"",
        ]
        
        for sf in sql_files:
            if sf.name == 'master_import.sql': continue
            master.append(f"-- ═══ {sf.name} ═══")
            master.append(f".read {sf.name}")
            master.append("")
        
        master.extend([
            f"COMMIT;",
            f"",
            f"-- ═══ POST-IMPORT VALIDATION ═══",
            f"SELECT 'Customers: ' || COUNT(*) FROM sale_customers;",
            f"SELECT 'Opportunities: ' || COUNT(*) FROM sale_opportunities;",
            f"SELECT 'Contract Milestones: ' || COUNT(*) FROM sale_contract_milestones;",
            f"SELECT 'Settlements: ' || COUNT(*) FROM sale_settlements;",
            f"SELECT 'NAS File Links: ' || COUNT(*) FROM sale_nas_file_links;",
            f"SELECT 'Quotation Revisions: ' || COUNT(*) FROM sale_quotation_revisions;",
            f"SELECT 'Digital Content: ' || COUNT(*) FROM sale_digital_content;",
            f"SELECT 'Import Log: ' || COUNT(*) FROM sale_import_log;",
            f"",
            f"-- ═══ INTEGRITY CHECKS ═══",
            f"SELECT 'Orphan opportunities (no customer): ' || COUNT(*) FROM sale_opportunities WHERE customer_id IS NOT NULL AND customer_id NOT IN (SELECT id FROM sale_customers);",
            f"SELECT 'Product groups: ' || GROUP_CONCAT(DISTINCT product_group) FROM sale_opportunities;",
            f"SELECT 'Stages: ' || GROUP_CONCAT(DISTINCT stage) FROM sale_opportunities;",
        ])
        
        with open(self.out / 'master_import.sql', 'w') as f:
            f.write('\n'.join(master))
        self.log(f"  → master_import.sql generated")

    def _write_sql(self, filename, lines, table_name):
        filepath = self.out / filename
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(f"-- {table_name}\n")
            f.write(f"-- Generated: {self.now}\n")
            f.write(f"-- Records: {len(lines)}\n\n")
            f.write('\n'.join(lines))
            f.write('\n')
        self.log(f"  Written: {filename} ({len(lines)} records)")

    def run(self):
        self.log("="*60)
        self.log("IBS HI SALE PLATFORM - DATABASE IMPORT PIPELINE")
        self.log("="*60)
        
        # Batch 1: Customers
        self.import_clients()
        
        # Batch 2: Pipeline (from IBSHI Khach hang tiem nang)
        self.import_pipeline()

        # Batch 2b: Active Contracts & Projects (from Sale Plan)
        self.import_active_contracts()

        # Batch 3: Invoice Status → Contract Milestones
        self.import_invoice_status()
        
        # Batch 4: Capital Recovery → Settlements
        self.import_thu_hoi_von()
        
        # Batch 5: VOGT Pipeline
        self.import_vogt_pipeline()
        
        # Batch 6: Quotation Files
        self.import_quotation_files()
        
        # Batch 7: Contract Files
        self.import_contracts()
        
        # Batch 8: MKT Content
        self.import_mkt_content()
        
        # Batch 9: Product Portfolio
        self.import_product_portfolio()
        
        # Audit
        self.generate_import_log()
        self.generate_master_script()
        
        # Summary
        self.log("")
        self.log("="*60)
        self.log("IMPORT COMPLETE - SUMMARY")
        self.log("="*60)
        for k, v in self.stats.items():
            self.log(f"  {k}: {v} records")
        total = sum(self.stats.values())
        self.log(f"  TOTAL: {total} records")
        self.log(f"  Output: {self.out}")
        self.log(f"  Run: sqlite3 sale.db < {self.out}/master_import.sql")

if __name__ == '__main__':
    sale_root = sys.argv[1] if len(sys.argv) > 1 else SALE_ROOT
    output_dir = sys.argv[2] if len(sys.argv) > 2 else OUTPUT_DIR
    
    importer = SalePlatformImporter(sale_root, output_dir)
    importer.run()
